import ast                                          #used for converting values
from email import encoders                          #modules for working with email messages, including creating, and parsing emails.
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
from playsound import playsound
from pyzbar.pyzbar import decode                    #decoding QR codes and other barcodes.
from PIL import Image, ImageTk                      #Python Imaging Library (PIL) is used for opening, manipulating, and saving images.
import customtkinter as ctk                          #creating custom Tkinter GUIs.
from tkinter import messagebox
from ultralytics import YOLO                        #YOLOv8
import supervision as sv                            #library for annotating and visualizing object detection results
import threading                                    #built-in Python module for managing concurrent threads
from datetime import datetime                       #module for working with dates and times
from openpyxl import Workbook, load_workbook        #Python library for reading and writing Excel files
from openpyxl.styles import Font, PatternFill
from openpyxl.drawing.image import Image as xImage
import pandas as pd                                 #data analysis and manipulation library
import smtplib                                      #module for sending email messages.
import cv2                                          #library for computer vision tasks like image processing and video analysis.
import csv                                          #module for reading and writing CSV files.
import sys                                          #module for system-specific parameters and functions.
import os                                           #module for interacting with the operating system
import sqlite3
import bcrypt
import io


class MultiCamApp:                                  #main app class
    def __init__(self, root):
        self.root = root
        self.root.title("Live Camera Feed")
        # self.root.geometry("1600x900")  # Set the window size
        self.root.attributes('-fullscreen', True)
        # Show login popup at the start
        self.operator_login()

        self.today = datetime.today()
        self.timeNow = datetime.now()
        date_formatted = self.today.strftime("%Y-%m-%d")
        # Create the "NewFolder" directory if it doesn't exist
        self.folder_path = os.path.join(os.path.expanduser("~\Documents"), "UniCheck")
        os.makedirs(self.folder_path, exist_ok=True)
        # Create the CSV file path
        self.csv_file_path = os.path.join(self.folder_path, f"UniCheck[{date_formatted}].csv")
        # Create the XLSX file path
        self.xlsx_file_path = os.path.join(self.folder_path, f"UniCheck_Report[{date_formatted}].xlsx")
        # Create txt file for system logging
        self.txt_file_path = os.path.join(self.folder_path, f"SysLog.txt")

        # Class imports
        self.detector = ObjectDetector("custom_model_version5.pt")
        self.qr_detector = QRDetector
        self.csv_handler = CSVHandler
        self.email_handler = EmailHandler()
    
        self.qr_data_dict= {}
        self.update_dict = {}
        self.remarks = ""
        self.hairStatus = ""
        self.email_sent = False

        # Flags to control camera feed
        self.cam1_running = False
        self.cam2_running = False
        self.cam1_frame = None  # Variable to store the current frame of Camera 1
        self.cam1_stored_frame = None
        self.sound_is_playing = False
        
        topframe = ctk.CTkFrame(self.root, fg_color="#4F6932")
        topframe.pack(fill="both", padx=20, pady=(20, 0))
        # Open the logo image
        logoPNG = Image.open('CCA-logo.png')
        width, height = logoPNG.size
        aspect_ratio = width / height

        # Determine the new dimensions while preserving the aspect ratio
        if width > height:
            new_width = 100
            new_height = int(new_width / aspect_ratio)
        else:
            new_height = 100
            new_width = int(new_height * aspect_ratio)

        # Resize the image to fit within the 100x100 canvas
        logoPNG = logoPNG.resize((new_width, new_height), Image.Resampling.LANCZOS)
        logoPNGimg = ImageTk.PhotoImage(logoPNG)
        self.root.logoPNGimg = logoPNGimg  # Prevent garbage collection

        # Create the canvas
        ccaLogo = ctk.CTkCanvas(topframe, width=100, height=100, highlightthickness=1, highlightbackground="black", background="white")
        ccaLogo.grid(row=0, column=0, sticky="nsew")

        # Center the image within the canvas
        x_offset = (100 - new_width) // 2
        y_offset = (100 - new_height) // 2
        ccaLogo.create_image(x_offset, y_offset, anchor="nw", image=logoPNGimg)

        # Add the "UniCheck" text next to the logo
        unicheck_label = ctk.CTkLabel(topframe, text="UniCheck", font=("Arial", 24, "bold"))
        unicheck_label.grid(row=0, column=1, sticky="w", padx=(0, 10))
        
        # Add the exit button on the rightmost side
        exit_button = ctk.CTkButton(topframe, text="Exit", command= self.on_close, fg_color="red", text_color="white")
        exit_button.grid(row=0, column=2, sticky="e", padx=(10, 0))

        # Adjust column weights to push the exit button to the rightmost side
        topframe.columnconfigure(0, weight=0)  # Logo column
        topframe.columnconfigure(1, weight=1)  # UniCheck text column
        topframe.columnconfigure(2, weight=0)  # Exit button column

        # Create a frame to hold the grid layout
        mainframe = ctk.CTkFrame(self.root)
        mainframe.pack(fill="both", expand=True, padx=20, pady=(0, 20))  # Make the frame fill the entire window

        # Configure the grid layout
        for i in range(8):
            mainframe.rowconfigure(i, weight=1, uniform=True)
            mainframe.columnconfigure(i, weight=1, uniform=True)

        default_image = Image.open("image.png")
        self.photo_image = ImageTk.PhotoImage(image=default_image)

        #Canvas_1 is where Camera 1 outputs
        self.canvas_1 = ctk.CTkCanvas(mainframe, highlightthickness = 1, highlightbackground = "black", background = "white")
        self.canvas_1.grid(row=0, column=0, rowspan=6, columnspan=6, sticky="nsew")
        
        #Canvas_2 is where Camera 2 outputs
        self.canvas_2 = ctk.CTkCanvas(mainframe, highlightthickness = 1, highlightbackground = "black", background = "white")
        self.canvas_2.grid(row=4, column=4, rowspan=2, columnspan=2, sticky="nsew")
        
        #Canvas_3 is where snapshot outputs
        self.canvas_3 = ctk.CTkCanvas(mainframe, highlightthickness = 1, highlightbackground = "black", background = "white")
        self.canvas_3.grid(row=0, column=6, rowspan=2, columnspan=2, sticky="nsew")    

        self.button_open_1 = ctk.CTkButton(mainframe, text="Open Cam 1", command= self.cam1Toggle)
        self.button_open_1.grid(column=0, row=6, rowspan=2, padx=50, pady=50, sticky="nsew")

        self.button_open_2 = ctk.CTkButton(mainframe, text="Open Cam 2", command= self.cam2Toggle)
        self.button_open_2.grid(column=1, row=6, rowspan=2, padx=50, pady=50, sticky="nsew")

        self.text_field = ctk.CTkTextbox(mainframe, activate_scrollbars=True, wrap='word', font=('Arial', 18))
        self.text_field.grid(column=2, row=6, columnspan=2, rowspan=2, sticky="nsew")
        self.text_field.configure(state="disabled")

        self.button_snaps = ctk.CTkButton(mainframe, text="View Recent Snaps", command= self.open_folder_captures)
        self.button_snaps.grid(column=4, row=6, padx=25, pady=25, sticky="nsew")

        self.button_logs = ctk.CTkButton(mainframe, text="View System Logs", command= self.open_system_logs)
        self.button_logs.grid(column=4, row=7, padx=25, pady=25, sticky="nsew")

        self.button_capture = ctk.CTkButton(mainframe, text="Capture", command= self.capture_cam1)
        self.button_capture.grid(column=5, row=6, padx=25, pady=25, sticky="nsew")

        self.button_generate_ticket = ctk.CTkButton(mainframe, text="Generate Ticket", command= self.on_generate_ticket_press)
        self.button_generate_ticket.grid(column=5, row=7, padx=25, pady=25, sticky="nsew")

        # student_IDPhoto = ctk.CTkFrame(mainframe)
        # student_IDPhoto.grid(column=6, row=2, columnspan=2, rowspan=4, sticky="nsew")

        self.student_IDPhoto = ctk.CTkCanvas(mainframe,width=250, height=250, highlightthickness = 1, highlightbackground = "black", background = "white")
        self.student_IDPhoto.grid(column=6, row=2, columnspan=2, rowspan=3)   

        student_info_field = ctk.CTkFrame(mainframe)
        student_info_field.grid(column=6, row=5, columnspan=2, rowspan=6, sticky="nsew")

        student_info_field.rowconfigure(0, uniform=True)
        student_info_field.rowconfigure(1, uniform=True)
        student_info_field.rowconfigure(2, uniform=True)
        student_info_field.rowconfigure(3, uniform=True)

        self.student_number = ctk.CTkLabel(student_info_field, text="Student No. : ", font=("Arial", 18))
        self.student_number.grid(row=0, sticky="w", ipadx= 10, ipady= 10)

        self.student_name = ctk.CTkLabel(student_info_field, text="Student Name : ", font=("Arial", 18))
        self.student_name.grid(row=1, sticky="w", ipadx= 10, ipady= 10)

        self.student_course = ctk.CTkLabel(student_info_field, text="Course-Section : ", font=("Arial", 18))
        self.student_course.grid(row=2, sticky="w", ipadx= 10,ipady= 10)

        self.student_gmail = ctk.CTkLabel(student_info_field, text="GMail : ", font=("Arial", 18))
        self.student_gmail.grid(row=3, sticky="w", ipadx= 10, ipady= 10)

        # Update all changes made to the layout
        self.root.update()

        # Global variable to store the PhotoImage object reference
        self.image_id_1 = self.canvas_1.create_image(self.canvas_1.winfo_width()//2, self.canvas_1.winfo_height()//2, anchor="center")
        self.image_id_2 = self.canvas_2.create_image(self.canvas_2.winfo_width()//2, self.canvas_2.winfo_height()//2, anchor="center")
        # self.image_id_3 = self.canvas_3.create_image(self.canvas_3.winfo_width()//2, self.canvas_3.winfo_height()//2, anchor="center")
    
    def cam1Toggle(self):
        if self.button_open_1._text == 'Open Cam 1':
                self.button_open_1.configure(text="Close Cam 1")
                self.start_cam1()
        else:
            self.button_open_1.configure(text="Open Cam 1")
            self.stop_cam1()
    
    def cam2Toggle(self):
        if self.button_open_2._text == 'Open Cam 2':
                self.button_open_2.configure(text="Close Cam 2")
                self.start_cam2()
        else:
            self.button_open_2.configure(text="Open Cam 2")
            self.stop_cam2()
    
    def start_cam1(self):
        if not self.cam1_running:
            self.cam1_running = True
            self.cam1_thread = threading.Thread(target=self.camera_feed, args=(1, self.canvas_1, "cam1"))
            # Enable or disable buttons based on camera status
            # self.button_open_1.configure(state="disabled")
            # self.button_close_1.configure(state="normal")
            self.cam1_thread.start()
            # self.append_to_terminal("Camera 1 enabled")

    def stop_cam1(self):
        self.cam1_running = False
        # self.button_open_1.configure(state="normal")
        # self.button_close_1.configure(state="disabled")
        # self.append_to_terminal("Camera 1 disabled")

    def start_cam2(self):
        if not self.cam2_running:
            self.cam2_running = True
            self.cam2_thread = threading.Thread(target=self.camera_feed, args=(0, self.canvas_2, "cam2"))
            # Enable or disable buttons based on camera status
            # self.button_open_2.configure(state="disabled")
            # self.button_close_2.configure(state="normal")
            self.cam2_thread.start()
            # self.append_to_terminal("Camera 2 enabled")

    def stop_cam2(self):
        self.cam2_running = False
        # self.button_open_2.configure(state="normal")
        # self.button_close_2.configure(state="disabled")
        # self.append_to_terminal("Camera 2 disabled")

    def camera_feed(self, cam_index, canvas, cam_name):
        cap = cv2.VideoCapture(cam_index, cv2.CAP_DSHOW)
        if cam_name == "cam1": 
            cap.set(cv2.CAP_PROP_FRAME_WIDTH, 1280)
            cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)
        if cam_name == "cam2": 
            cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
            cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 360)
        while getattr(self, f'{cam_name}_running'):
            ret, frame = cap.read()
            if ret:
                frame = cv2.flip(frame, 1)
                if cam_name == "cam1":                      #Limit the uniform detection to camera 1 only
                    frame, self.remarks = self.detector.detect(frame)  # Detect objects in the frame
                    if self.remarks != "":
                        self.cam1_stored_frame = frame
                        if (self.remarks == 'Not_In_Uniform ') or (self.remarks == "Not_In_Uniform Hair_Allowed") or (self.remarks == "In_Uniform Hair_Not_Allowed") or (self.hairStatus ==  " Hair_Not_Allowed"):
                            self.play_notification_sound()
                    if (bool(self.qr_data_dict)):
                        self.remarks = self.remarks
                if cam_name == "cam2":                      #Limit the QR detection to camera 2 only
                    qr_data = self.qr_detector.read_qr_code(frame=frame)  # Detect QR in the frame
                    if qr_data is not None:
                        qr_data_str = ast.literal_eval(qr_data)                 #Convert value into string
                        self.qr_data_dict = ast.literal_eval(qr_data_str)       #Convert string into dictionary
                        self.appendStudentInfo(studNo = self.qr_data_dict["StudentID"], studName=self.qr_data_dict["StudentName"], studCourse=self.qr_data_dict["Course-Section"], studMail=self.qr_data_dict["Gmail"])
                        # self.append_to_terminal("QR Data : "+output_string)

                frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                img = Image.fromarray(frame)
                if cam_name == "cam1":
                    img = img.resize((canvas.winfo_width(), canvas.winfo_height()), Image.NEAREST)
                imgtk = ImageTk.PhotoImage(image=img)

                # Store the frame for respective camera if it's running
                if cam_name == "cam1":
                    canvas.itemconfig(self.image_id_1, image=imgtk)
                    canvas.image = imgtk 
                    self.cam1_frame = frame
                if cam_name == "cam2":
                    canvas.itemconfig(self.image_id_2, image=imgtk)
                    canvas.image = imgtk 
 
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break
            # print(threading.active_count(),threading.enumerate(),self.cam1_thread.is_alive(),self.cam2_thread.is_alive())
        cap.release()
        canvas.itemconfig(self.image_id_1, image=self.photo_image)
        canvas.image = imgtk 

    def capture_cam1(self):
        if self.qr_data_dict["StudentID"] is None:
            self.append_to_terminal("No QR Detected!")
            return
        if self.cam1_frame is not None:
            # Convert the stored frame from Camera 1 to display it
            
            img = Image.fromarray(self.cam1_frame)
            imgtk = ImageTk.PhotoImage(image=img)
            self.canvas_3.create_image(self.canvas_3.winfo_width()//2, self.canvas_3.winfo_height()//2, anchor="center", image=imgtk)
            self.canvas_3.image = imgtk

            #Merge QR data and Captured Image
            studentID = self.qr_data_dict["StudentID"]
            studentEmail = self.qr_data_dict["Gmail"]
            image_path = self.detector.save_image_to_folder(image = self.cam1_stored_frame, student_id=studentID)

            #Add new column to Dictionary
            self.qr_data_dict['Remarks'] = self.remarks
            self.update_dict['Time_of_entry']  = self.timeNow.strftime("%H:%M:%S")
            self.update_dict.update(self.qr_data_dict)

            #Send email only if not in uniform or hair not allowed
            if (self.remarks == 'Not_In_Uniform ') or (self.remarks == "Not_In_Uniform Hair_Allowed") or (self.remarks == "In_Uniform Hair_Not_Allowed") or (self.hairStatus ==  " Hair_Not_Allowed"):
                self.email_handler.send_email(receiver_email=studentEmail, image_path=image_path)
                self.email_sent = True
                self.append_to_terminal(f"Email sent to: {studentID}")
            else:
                self.email_sent = False

            #csv and log
            self.csv_handler.log_data_to_csv(self.update_dict, image_path, self.csv_file_path)
            self.csv_handler.generate_xlsx_with_images(self.csv_file_path, self.xlsx_file_path)
            self.csv_handler.format_xlsx(self.xlsx_file_path)
            self.append_to_terminal(f"Logged successfully: {studentID}")          
            self.cam1_frame = None
            self.start_cam2()
        else:
            self.append_to_terminal("No QR and Image Capture!")

    def dbImageSearch(self, studentID):
        # Fetch image path from the database
        conn = sqlite3.connect(".dbDoNotDelete\\unicheck.db")
        cursor = conn.cursor()
    
        cursor.execute("SELECT id_image_path FROM students WHERE student_id = ?", (studentID,))
        image_path = cursor.fetchone()[0]  # Get the image path

        # Display the image using PIL
        image = Image.open(image_path)
        width, height = image.size
        aspect_ratio = width / height

        # Determine the new dimensions while preserving the aspect ratio
        if width > height:
            new_width = 250
            new_height = int(new_width / aspect_ratio)
        else:
            new_height = 250
            new_width = int(new_height * aspect_ratio)

        # Resize the image to fit within the 100x100 canvas
        image = image.resize((new_width, new_height), Image.Resampling.LANCZOS)
        image = ImageTk.PhotoImage(image)
        self.root.student_image = image  # Prevent garbage collection

        # Center the image within the canvas
        x_offset = (250 - new_width) // 2
        y_offset = (250 - new_height) // 2
        self.student_IDPhoto.create_image(x_offset, y_offset, anchor="nw", image=image)
        # image.show()  # Opens the image in the default image viewer
        conn.close()

    def appendStudentInfo(self, studNo, studName, studCourse, studMail):
        self.dbImageSearch(studNo)
        self.student_number.configure(text="Student No. : " + studNo)
        self.student_name.configure(text="Student Name : " + studName)
        self.student_course.configure(text="Course-Section : " + studCourse)
        self.student_gmail.configure(text="GMail : " + studMail)

    def append_to_terminal(self, text):
        # Enable text_area to append new text
        self.text_field.configure(state="normal")
        # Insert the new text at the end
        self.text_field.insert(ctk.END, text + "\n")
        self.append_system_logs(text_to_append=self.timeNow.strftime("%m\%d\%y %H:%M:%S")+" "+text)
        # Scroll to the bottom
        self.text_field.yview(ctk.END)
        # Disable editing again to make it read-only
        self.text_field.configure(state=ctk.DISABLED)
    
    def on_close(self):
        top_level = ctk.CTkToplevel(root)
        width=300
        height=200
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        x = (screen_width/2) - (width/2)
        y = (screen_height/2) - (height/2)
        top_level.geometry('%dx%d+%d+%d' % (width, height, x, y))
        top_level.wm_transient(root)

        if (self.cam1_running or self.cam2_running):
            top_level.title("Warning")

            label = ctk.CTkLabel(top_level, text="Close runnning cameras!")
            label.pack(pady=30)

            ok_button = ctk.CTkButton(top_level, text="Ok", command=lambda: [top_level.destroy()])
            ok_button.pack(side="bottom", padx=10, pady=10)

            top_level.mainloop()
        else:
            top_level.title("Confirmation")

            label = ctk.CTkLabel(top_level, text="Are you sure you want to quit?")
            label.pack(pady=30)

            yes_button = ctk.CTkButton(top_level, text="Yes", command=lambda: [root.destroy()])
            yes_button.pack(side="left", padx=10, pady=10)

            no_button = ctk.CTkButton(top_level, text="No", command=top_level.destroy)
            no_button.pack(side="right", padx=10, pady=10)

            top_level.mainloop()
    
    def operator_login(self):
        # Create the login popup
        login_popup = ctk.CTkToplevel(self.root)
        login_popup.title("Operator Login")
        login_popup.geometry("300x200")
        login_popup.grab_set()  # Prevent interaction with the main window
        login_popup.protocol("WM_DELETE_WINDOW", lambda: [root.destroy()])  # Disable close button

        # Center the popup
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width // 2) - (300 // 2)
        y = (screen_height // 2) - (200 // 2)
        login_popup.geometry(f"300x200+{x}+{y}")

        # Login UI
        ctk.CTkLabel(login_popup, text="Username or Email").pack(pady=5)
        username_entry = ctk.CTkEntry(login_popup)
        username_entry.pack(pady=5)

        ctk.CTkLabel(login_popup, text="Password").pack(pady=5)
        password_entry = ctk.CTkEntry(login_popup, show="*")
        password_entry.pack(pady=5)

        login_button = ctk.CTkButton(
            login_popup,
            text="Login",
            command=lambda: self.validate_login(
                username_entry.get(), password_entry.get(), login_popup
            ),
        )
        login_button.pack(pady=10)

    def validate_login(self, username, password, login_popup):
        # Connect to the database to validate credentials
        conn = sqlite3.connect(".dbDoNotDelete\\unicheck.db")
        cursor = conn.cursor()

        # Fetch the stored password hash for the username/email
        cursor.execute(
            "SELECT password_hash FROM operators WHERE username_or_email = ?",
            (username,),
        )
        result = cursor.fetchone()
        conn.close()

        if result:
            stored_hash = result[0]  # This is already a bytes object
            if bcrypt.checkpw(password.encode("utf-8"), stored_hash):
                # Login successful
                messagebox.showinfo("Login", "Login successful!")
                login_popup.destroy()
            else:
                # Invalid password
                messagebox.showerror("Login Failed", "Invalid password. Try again.")
        else:
            # Username not found
            messagebox.showerror("Login Failed", "User not found. Try again.")

    def on_generate_ticket_press(self):
        if(self.email_sent):
            self.append_to_terminal("Email already sent!")
            return
        if (self.qr_data_dict.get("Gmail")):
            if (self.cam1_stored_frame is not None):
                self.email_handler.send_email(self.qr_data_dict["Gmail"],self.cam1_stored_frame)
            else:
                self.append_to_terminal("No image captured!")
        else:
            self.append_to_terminal("Email not found!")
    
    def open_folder_captures(self):
        folder_path = os.path.join(os.path.expanduser("~\\Documents\\UniCheck"), "Captured_Images")  # Replace with the desired folder path
        os.startfile(folder_path)

    def append_system_logs(self, text_to_append):   
        with open(self.txt_file_path, 'a') as file:
            file.write(text_to_append + '\n')

    def open_system_logs(self):
        folder_path = self.txt_file_path
        os.startfile(folder_path)

    def play_notification_sound(self):
        if not self.sound_is_playing:
            self.sound_is_playing = True
            threading.Thread(target=self.play_and_reset_flag, daemon=True).start()

    def play_and_reset_flag(self):
        playsound("beeps_notification.mp3")
        time.sleep(2)
        self.sound_is_playing = False


class ObjectDetector:                               #object detection and annotation
    def __init__(self, model_path):
        self.model = YOLO(model_path)  # Load your custom object detection model
        # self.color_lookup = {
        #     "HairAllowed": sv.Color(0, 255, 0),
        #     "HairNotAllowed": sv.Color(255, 0, 0),
        #     "InUniform": sv.Color(0, 255, 0),
        #     "NotInUniform": sv.Color(255, 0, 0),
        #     "PE/NSTP":sv.Color(0, 255, 0)
        # }

    def detect(self, frame):
        self.label_annotator1 = sv.LabelAnnotator(text_position=sv.Position.TOP_CENTER,text_scale=0.5,text_thickness=1,text_padding=10,color=sv.Color(255, 0, 0))
        self.box_annotator1 = sv.BoxCornerAnnotator(thickness=2,color=sv.Color(255, 0, 0))      #Red

        self.label_annotator2 = sv.LabelAnnotator(text_position=sv.Position.TOP_CENTER,text_scale=0.5,text_thickness=1,text_padding=10,color=sv.Color(0, 255, 0))
        self.box_annotator2 = sv.BoxCornerAnnotator(thickness=2,color=sv.Color(0, 255, 0))      #Green

        self.label_annotator = sv.LabelAnnotator(text_position=sv.Position.TOP_CENTER,text_scale=0.5,text_thickness=1,text_padding=10)
        self.box_annotator = sv.BoxCornerAnnotator(thickness=2)   

        result = self.model(frame, conf=0.50, show_labels=True, show_conf=True, verbose=False)[0]  # Perform detection
        detections = sv.Detections.from_ultralytics(result)  # Convert results to sv.Detections format
        
        #Assign new name to classes
        # mapped_class_names = [self.class_name_map.get(class_name, class_name) for class_name in detections['class_name']]

        #Annotate the frame with bounding boxes and labels
        labels = [f"{class_name} {confidence:.2f}" for class_name, confidence in zip(detections['class_name'], detections.confidence)]
        
        hair_status = ""
        remark = ""
        
        for class_name in detections['class_name']:
            label_annotator = sv.LabelAnnotator(text_position=sv.Position.TOP_CENTER,text_scale=0.5,text_thickness=1,text_padding=10)
            box_annotator = sv.BoxCornerAnnotator(thickness=2)
            
            if class_name == "HairNotAllowed":
                hair_status = "Hair_Not_Allowed"
                frame = label_annotator.annotate(scene=frame, detections=detections, labels=labels)
                frame = box_annotator.annotate(scene=frame, detections=detections)
            if class_name == "HairAllowed":
                hair_status = "Hair_Allowed"
                frame = label_annotator.annotate(scene=frame, detections=detections, labels=labels)
                frame = box_annotator.annotate(scene=frame, detections=detections)
            frame2 = frame
            if class_name == "InUniform":
                frame2 = label_annotator.annotate(scene=frame2, detections=detections, labels=labels)
                frame2 = box_annotator.annotate(scene=frame2, detections=detections)
                remark = "In_Uniform " + hair_status
                return frame2 , remark
            if class_name == "NotInUniform":
                frame2 = label_annotator.annotate(scene=frame2, detections=detections, labels=labels)
                frame2 = box_annotator.annotate(scene=frame2, detections=detections)
                remark = "Not_In_Uniform " + hair_status
                return frame2 , remark
            if class_name == "PE/NSTP":
                frame2 = label_annotator.annotate(scene=frame2, detections=detections, labels=labels)
                frame2 = box_annotator.annotate(scene=frame2, detections=detections)
                remark = "PE/NSTP " + hair_status
                return frame2 , remark

        return frame, ""

    def save_image_to_folder(self, image, student_id):
        # Create the folder if it doesn't exist
        today = datetime.now()
        date_formatted = today.strftime("%m%d%y_%H%M%S")
        folder_path = os.path.join(os.path.expanduser("~\\Documents\\UniCheck"), "Captured_Images")
        os.makedirs(folder_path, exist_ok=True)

        # Create the file path using the student ID or timestamp as the filename
        file_name = f"{date_formatted}_{student_id}.jpg"  # You can include timestamp or other identifiers
        file_path = os.path.join(folder_path, file_name)

        # Save the image
        cv2.imwrite(file_path, image)
        # Return the file path for future reference
        return file_path


class QRDetector:               #class for handling qr detection
    def read_qr_code(frame):
        # Convert the frame to grayscale for better QR code detection
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        # Detect QR codes in the frame
        qrcodes = decode(gray)

        if qrcodes:
            # Extract data from decoded object
            data = qrcodes[0].data.decode("utf-8")
            # print("Decoded Data:", data)
            app.stop_cam2()
            # app.start_cam1()
            return data
                 

class CSVHandler:               #class for handling csv and generating xlsx reports
    def log_data_to_csv(student_data, image_path, csv_file_path):
        file_exists = os.path.isfile(csv_file_path)
        # student_data['Remarks'] = remarks

        # Add image path to the student data
        student_data['Capture'] = image_path
        # print(student_data)
        # Write the student data to the CSV file
        with open(csv_file_path, 'a', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            # Write headers only if file doesn't exist
            if not file_exists:
                writer.writerow(["Time_of_entry","StudentID", "StudentName", "Course-Section", "Gmail", "Remarks", "Capture"])
            writer.writerow([student_data['Time_of_entry'], student_data['StudentID'], student_data['StudentName'], 
                            student_data['Course-Section'],student_data['Gmail'], student_data['Remarks'], student_data['Capture']])

    def read_csv(csv_file_path):
        # Read CSV into pandas DataFrame
        df = pd.read_csv(csv_file_path)
        print(df.head())  # Just to view the first few rows of data
        return df

    def appendFiles(data, csv_file_path):
        with open(csv_file_path, 'a', newline='', encoding="utf-8") as csvfile:                 #Writing a csv file with UTF-8 encoding
            writer = csv.writer(csvfile)                                                        #New Instance of csv.writer
            writer.writerow(["Time_of_entry","StudentID", "StudentName", "Course-Section", "Gmail", "Remarks", "Capture"])                   #Assigning columns
            for item in data:
                writer.writerow([item['Time_of_entry'], item['StudentID'], item['StudentName'], item['Course-Section'], item['Gmail'], item['Remarks'], item['Capture']])

    def format_xlsx(xlsx_file_path):
        try:
            # Load the workbook and select the sheet
            wb = load_workbook(xlsx_file_path)
            ws = wb['Attendance']

            # Apply formatting to the header
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

            for cell in ws[1]:  # First row is the header
                cell.font = header_font
                cell.fill = header_fill

            # Save changes
            wb.save(xlsx_file_path)
            # app.append_to_terminal(f"Formatted XLSX saved to: {xlsx_file_path}")
        except PermissionError:
            app.append_to_terminal(f"Error: The file '{xlsx_file_path}' is currently open. Please close it and try again.")
    
    def generate_xlsx_with_images(csv_file_path, xlsx_file_path):
        try:
            # Read the CSV file
            df = pd.read_csv(csv_file_path)

            # Write the DataFrame to an Excel file
            with pd.ExcelWriter(xlsx_file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Attendance', index=False)
                
            # Create a new workbook
            wb = load_workbook(xlsx_file_path)
            ws = wb.active

            # Write headers
            for col_num, column in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_num, value=column)

            # Write data and add images
            for index, row in df.iterrows():
                for col_num, value in enumerate(row, 1):
                    ws.cell(row=index+2, column=col_num, value=value)

                # Insert images based on the image path stored in the CSV
                if 'Capture' in df.columns and row['Capture']:  # Assuming 'Capture' column has the image path
                    img_path = row['Capture']
                    if os.path.exists(img_path):  # Ensure the image exists
                        img = xImage(img_path)
                        # img.width = 100  # Resize image if necessary
                        # img.height = 100
                        # img_width, img_height = img.width, img.height
                        # print (img_width,img_height)
                        # # Calculate the cell dimensions (In Excel, the default row height is 15 (20 pixels), and the column width is 8.43 (64 pixels))
                        # column_width = img_width / 7  # Set column width based on image width
                        # row_height = img_height / 20  # Set row height based on image height
                        # # Set the column width and row height for the cell
                        ws.column_dimensions['G'].width = 90.71
                        ws.row_dimensions[index+2].height = 360

                        ws.add_image(img, f"G{index+2}")  # Place in the Capture column (e.g., column G)

            # Save the workbook
            wb.save(xlsx_file_path)
        except PermissionError:
            app.append_to_terminal(f"Error: The file '{xlsx_file_path}' is currently open. Please close it and try again.")
     

class EmailHandler:             #creating and sending emails
    def __init__(self):
            # Example usage
        self.sender_email = 'sample.sender@email.com'
        self.subject = 'Test Email'
        self.body = "We detected a violation for further information regarding this matter, kindly proceed to Student Affair and Services Office (SASO)."

    def send_email(self, receiver_email, image_path):
        def email_thread():
            # Compose the email message
            msg = MIMEMultipart()
            msg['From'] = self.sender_email
            msg['To'] = receiver_email
            msg['Subject'] = self.subject
            msg.attach(MIMEText(self.body, 'plain'))

            # Attach the image
            try:
                with open(image_path, 'rb') as img_file:
                    img = MIMEBase('application', 'octet-stream')
                    img.set_payload(img_file.read())
                    encoders.encode_base64(img)
                    img.add_header('Content-Disposition', 'attachment', filename='image.png')
                    msg.attach(img)
            except Exception as e:
                app.append_to_terminal(f"Error attaching image: {e}")
                return

            # Send the email using SMTP
            try:
                with smtplib.SMTP('smtp.gmail.com', 587) as smtp_server:
                    smtp_server.starttls()
                    smtp_server.login(self.sender_email, 'app pas sword')  # Use your app password here
                    smtp_server.send_message(msg)
                app.append_to_terminal("Email sent successfully!")
            except Exception as e:
                app.append_to_terminal(f"Failed to send email: {e}")
        # Start the email sending process in a separate thread
        threading.Thread(target=email_thread, daemon=True).start()


if __name__ == '__main__':          #MainLoop
    ctk.set_appearance_mode("light")
    ctk.set_default_color_theme("green")
    root = ctk.CTk()
    app = MultiCamApp(root)
    # app.start_cam1()
    # app.start_cam2()
    root.protocol("WM_DELETE_WINDOW", app.on_close)
    root.mainloop()
    sys.exit()
