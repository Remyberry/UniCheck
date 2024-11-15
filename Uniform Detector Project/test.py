from datetime import datetime
import cv2
import os
from openpyxl import Workbook, load_workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.drawing.image import Image as xImage
import pandas as pd
import csv

def save_image_to_folder(image, student_id):
    # Create the folder if it doesn't exist
    today = datetime.now()
    date_formatted = today.strftime("%y%m%d_%H%M%S")
    folder_path = os.path.join(os.path.expanduser("~\\Documents\\UniCheck"), "Captured_Images")
    os.makedirs(folder_path, exist_ok=True)

    # Create the file path using the student ID or timestamp as the filename
    file_name = f"{date_formatted}_{student_id}.jpg"  # You can include timestamp or other identifiers
    file_path = os.path.join(folder_path, file_name)

    # Save the image
    cv2.imwrite(file_path, image)
    
    # Return the file path for future reference
    return file_path

def log_data_to_csv(student_data, image_path, csv_file_path):

    file_exists = os.path.isfile(csv_file_path)
    # Add image path to the student data
    student_data['Capture'] = image_path

    # Write the student data to the CSV file
    with open(csv_file_path, 'a', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
         # Write headers only if file doesn't exist
        if not file_exists:
            writer.writerow(["StudentID", "StudentName", "Course-Section", "Email", "Remarks", "Capture"])
        writer.writerow([student_data['StudentID'], student_data['StudentName'], 
                         student_data['Course-Section'],student_data['Email'], student_data['Remarks'], student_data['Capture']])

def read_csv(csv_file_path):
    # Read CSV into pandas DataFrame
    df = pd.read_csv(csv_file_path)
    print(df.head())  # Just to view the first few rows of data
    return df

# def csv_to_xlsx(csv_file_path, xlsx_file_path):
#     # Read CSV data into pandas DataFrame
#     df = pd.read_csv(csv_file_path)

#     # Write the DataFrame to an Excel file
#     with pd.ExcelWriter(xlsx_file_path, engine='openpyxl') as writer:
#         df.to_excel(writer, sheet_name='Attendance', index=False)

#     print(f"XLSX file saved to: {xlsx_file_path}")

def format_xlsx(xlsx_file_path):
    try:
        # Load the workbook and select the sheet
        wb = load_workbook(xlsx_file_path)
        ws = wb['Attendance']

        # Apply formatting to the header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        for cell in ws[1]:  # First row is the header
            cell.font = header_font
            cell.fill = header_fill

        # Optionally, add conditional formatting (e.g., color rows based on the "Remarks" column)
        # Save changes
        wb.save(xlsx_file_path)
        print(f"Formatted XLSX saved to: {xlsx_file_path}")
    except PermissionError:
        print(f"Error: The file '{xlsx_file_path}' is currently open. Please close it and try again.")

def generate_xlsx_with_images(csv_file_path, xlsx_file_path):
    try:
        # Read the CSV file
        df = pd.read_csv(csv_file_path)

        # Write the DataFrame to an Excel file
        with pd.ExcelWriter(xlsx_file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Attendance', index=False)
            
        # Create a new workbook
        wb = load_workbook(xlsx_file_path)
        ws = wb.active

        # Write headers
        for col_num, column in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_num, value=column)

        # Write data and add images
        for index, row in df.iterrows():
            for col_num, value in enumerate(row, 1):
                ws.cell(row=index+2, column=col_num, value=value)

            # Insert images based on the image path stored in the CSV
            if 'Capture' in df.columns and row['Capture']:  # Assuming 'Capture' column has the image path
                img_path = row['Capture']
                if os.path.exists(img_path):  # Ensure the image exists
                    img = xImage(img_path)
                    # img.width = 100  # Resize image if necessary
                    # img.height = 100
                    # img_width, img_height = img.width, img.height
                    # print (img_width,img_height)
                    # # Calculate the cell dimensions (In Excel, the default row height is 15 (20 pixels), and the column width is 8.43 (64 pixels))
                    # column_width = img_width / 7  # Set column width based on image width
                    # row_height = img_height / 20  # Set row height based on image height
                    # # Set the column width and row height for the cell
                    ws.column_dimensions['F'].width = 90.71
                    ws.row_dimensions[index+2].height = 360

                    ws.add_image(img, f"F{index+2}")  # Place in the Capture column (e.g., column F)

        # Save the workbook
        wb.save(xlsx_file_path)
    except PermissionError:
        print(f"Error: The file '{xlsx_file_path}' is currently open. Please close it and try again.")

# Example usage
student_data = {
    'StudentID': '12345', 
    'StudentName': 'John Doe', 
    'Course-Section': 'CS-101', 
    'Email': 'blahblah@gmail.com',
    'Remarks': 'In Uniform'
}

today = datetime.today()
date_formatted = today.strftime("%Y-%m-%d")
folder_path = os.path.join(os.path.expanduser("~\Documents"), "UniCheck")
os.makedirs(folder_path, exist_ok=True)

# Create the CSV file path
csv_file_path = os.path.join(folder_path, f"UniCheck[{date_formatted}].csv")

# Create the XLSX file path
xlsx_file_path = os.path.join(folder_path, f"UniCheck_Report[{date_formatted}].xlsx")

# csv_file_path = "attendance.csv"
# Example usage
cap = cv2.VideoCapture(1, cv2.CAP_DSHOW)
ret, frame = cap.read()
if ret:
    image_path = save_image_to_folder(frame, student_data['StudentID'])             #sample student id
    print(f"Image saved at: {image_path}")
cap.release()

log_data_to_csv(student_data, image_path, csv_file_path)

# df = read_csv(csv_file_path)

# Example of converting CSV to XLSX
# xlsx_file_path = "attendance_report.xlsx"
# csv_to_xlsx(csv_file_path, xlsx_file_path)

# Example usage
generate_xlsx_with_images(csv_file_path, xlsx_file_path)
# Apply formatting to the previously created XLSX
format_xlsx(xlsx_file_path)

