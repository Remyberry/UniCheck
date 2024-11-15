import pandas as pd
import csv
import os
from datetime import datetime
import logging

def append_to_csv(data, csv_file_path):
    # Check if the CSV file exists
    file_exists = os.path.isfile(csv_file_path)

    # Open CSV file for appending data
    with open(csv_file_path, 'a', newline='', encoding="utf-8") as csvfile:
        writer = csv.writer(csvfile)
        
        # Write headers only if file doesn't exist
        if not file_exists:
            writer.writerow(["StudentID", "StudentName", "Course-Section", "Email", "Remarks", "Capture"])

        # Append the new row of data
        writer.writerow(data)

def read_csv(csv_file_path):
    # Read CSV into pandas DataFrame
    df = pd.read_csv(csv_file_path)
    print(df.head())  # Just to view the first few rows of data
    return df


# Example usage of appending data
csv_file_path = "attendance.csv"
data = ['12345', 'John Doe', 'CS-101', 'john@example.com', 'In Uniform', 'path/to/snapshot.jpg']
append_to_csv(data, csv_file_path)

# Reading data back from CSV
df = read_csv(csv_file_path)


def csv_to_xlsx(csv_file_path, xlsx_file_path):
    # Read CSV data into pandas DataFrame
    df = pd.read_csv(csv_file_path)

    # Write the DataFrame to an Excel file
    with pd.ExcelWriter(xlsx_file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Attendance', index=False)

    print(f"XLSX file saved to: {xlsx_file_path}")

# Example of converting CSV to XLSX
xlsx_file_path = "attendance_report.xlsx"
csv_to_xlsx(csv_file_path, xlsx_file_path)


from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

def format_xlsx(xlsx_file_path):
    # Load the workbook and select the sheet
    wb = load_workbook(xlsx_file_path)
    ws = wb['Attendance']

    # Apply formatting to the header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for cell in ws[1]:  # First row is the header
        cell.font = header_font
        cell.fill = header_fill

    # Optionally, add conditional formatting (e.g., color rows based on the "Remarks" column)
    # Save changes
    wb.save(xlsx_file_path)
    print(f"Formatted XLSX saved to: {xlsx_file_path}")

# Apply formatting to the previously created XLSX
format_xlsx(xlsx_file_path)
